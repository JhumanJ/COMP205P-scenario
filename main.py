#Imports
import ast, math
from openpyxl import Workbook


#Open File
def calc_dist(my_list):
    my_sum = 0
    for i in range(len(my_list)-1):
        distance = math.sqrt( ((my_list[i][0]-my_list[i+1][0])**2)+((my_list[i][1]-my_list[i+1][1])**2) )
        my_sum = my_sum + distance
    return my_sum

def main():
    f = open('robots.mat.txt','r')

    # Excel code
    # column_number = 1
    # wb = Workbook()
    # ws = wb.active
    # ws.title = "Data"

    for i in range (1,31):
        text = f.readline().partition("\n")[0]
        text = text.partition(": ")[2]
        text = text.partition("#")
        #Read file
        points = text[0]
        temp_obstacles = text[2].split(";")

        #Parse Points
        try:
            points = list(ast.literal_eval(points))
        except(SyntaxError,ValueError):
            pass

        # Parse obstacles
        obstacles = []
        for item in temp_obstacles:
            temp_item=()
            try:
                temp_item = list(ast.literal_eval(item))
            except(SyntaxError,ValueError):
                pass
            if temp_item != ():
                obstacles.append(temp_item)


        # Display info
        print('\nGraph '+str(i)+':')
        # print("Points ("+str(len(points))+"): " + str(points))
        # print("Obstacles ("+str(len(obstacles))+"): " + str(obstacles))
        # print("MATRIX:\n")


        # Matrix of paths
        paths = []

        for i in range(0,len(points)):
            tempList = []
            for j in range(0,len(points)):
                if i == j:
                    tempList.append([()])
                else:
                    tempTuple = [points[i],points[j]]
                    tempList.append(tempTuple)
            paths.append(tempList)
            # print(tempList)

        # Base case
        # if len(obstacles)==0:

    #     Excel Code
    #     my_index = 1
    #     ws.cell(row=my_index, column=column_number, value="Point X")
    #     ws.cell(row=my_index, column=column_number+1, value="Point Y")
    #     my_index = my_index + 1
    #
    #     for point in points:
    #         ws.cell(row=my_index, column=column_number, value=float(point[0]))
    #         ws.cell(row=my_index, column=column_number+1, value=float(point[1]))
    #         my_index = my_index + 1
    #
    #     ws.cell(row=my_index, column=column_number, value="Obstacles X")
    #     ws.cell(row=my_index, column=column_number+1, value="Obstacles Y")
    #     my_index = my_index + 1
    #
    #     for item in obstacles:
    #         for point in item:
    #             ws.cell(row=my_index, column=column_number, value=float(point[0]))
    #             ws.cell(row=my_index, column=column_number+1, value=float(point[1]))
    #
    #             my_index = my_index + 1
    #
    #         my_index = my_index + 1
    #
    #     column_number = column_number + 2
    #
    # Excel code
    # wb.save('data.xlsx')

main()
